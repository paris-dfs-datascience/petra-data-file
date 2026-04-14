"""Import audit rules from an Excel workbook into rules/rules.json.

Usage:
    python scripts/import_rules_from_excel.py [path/to/rules.xlsx]

Defaults to ~/Downloads/Audit Tool Rules.xlsx if no path is given.
Existing rule IDs are preserved; only new IDs are appended.
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl is required. Install with: pip install openpyxl\n")
    sys.exit(1)


REPO_ROOT = Path(__file__).resolve().parents[1]
RULES_JSON_PATH = REPO_ROOT / "rules" / "rules.json"
DEFAULT_XLSX = Path.home() / "Downloads" / "Audit Tool Rules.xlsx"

VISION_HINT_METHODS = {"visual_or_hint"}


def humanize_id(rule_id: str) -> str:
    """FMT-DOLLAR-SIGNS -> 'Dollar Sign Placement' style fallback name."""
    parts = rule_id.split("-")
    if len(parts) > 1:
        parts = parts[1:]
    return " ".join(p.capitalize() for p in parts)


def derive_query(description: str | None, pass_criteria: str | None, fail_criteria: str | None) -> str:
    pieces = []
    if description:
        pieces.append(description.strip())
    if pass_criteria:
        pieces.append(f"Pass criteria: {pass_criteria.strip()}")
    if fail_criteria:
        pieces.append(f"Fail criteria: {fail_criteria.strip()}")
    return "\n\n".join(pieces) if pieces else ""


def coerce_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip().lower() in {"true", "yes", "y", "1"}
    return bool(value)


def coerce_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def row_to_rule(row: dict) -> dict | None:
    rule_id = (row.get("Rule ID") or "").strip()
    if not rule_id:
        return None

    description = (row.get("Description") or "").strip() or None
    pass_criteria = (row.get("Pass Criteria") or "").strip() or None
    fail_criteria = (row.get("Fail Criteria") or "").strip() or None
    check_method = (row.get("Check Method") or "").strip() or None

    analysis_type = "vision" if check_method in VISION_HINT_METHODS else "text"

    return {
        "id": rule_id,
        "name": humanize_id(rule_id),
        "analysis_type": analysis_type,
        "query": derive_query(description, pass_criteria, fail_criteria),
        "description": description,
        "acceptance_criteria": pass_criteria,
        "severity": (row.get("Severity") or "").strip() or None,
        "group": (row.get("Group") or "").strip() or None,
        "section": (row.get("Section") or "").strip() or None,
        "bypassable": coerce_bool(row.get("Bypassable")),
        "tolerance": coerce_float(row.get("Tolerance")),
        "check_method": check_method,
        "steps": (row.get("Steps") or "").strip() or None,
        "pass_criteria": pass_criteria,
        "fail_criteria": fail_criteria,
        "action_if_fail": (row.get("Action if Fail") or "").strip() or None,
        "rationale": (row.get("Rationale") or "").strip() or None,
    }


def read_excel_rules(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["rules"] if "rules" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    rules = []
    for raw in rows[1:]:
        if not any(raw):
            continue
        record = dict(zip(headers, raw))
        rule = row_to_rule(record)
        if rule:
            rules.append(rule)
    return rules


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        sys.stderr.write(f"Excel file not found: {xlsx_path}\n")
        return 1

    new_rules = read_excel_rules(xlsx_path)
    print(f"Read {len(new_rules)} rules from {xlsx_path}")

    existing = {"rules": []}
    if RULES_JSON_PATH.exists():
        existing = json.loads(RULES_JSON_PATH.read_text())

    existing_ids = {r["id"] for r in existing.get("rules", [])}
    added = []
    skipped = []
    for rule in new_rules:
        if rule["id"] in existing_ids:
            skipped.append(rule["id"])
            continue
        existing.setdefault("rules", []).append(rule)
        added.append(rule["id"])

    RULES_JSON_PATH.write_text(json.dumps(existing, indent=2, ensure_ascii=False) + "\n")

    print(f"Added {len(added)} rules; skipped {len(skipped)} (already present).")
    if skipped:
        print("Skipped:", ", ".join(skipped))
    print(f"Wrote {RULES_JSON_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
